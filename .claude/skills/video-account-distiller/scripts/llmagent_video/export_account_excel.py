# coding=utf-8
import argparse
import json
import os
import sys

PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def parse_args():
    parser = argparse.ArgumentParser(description="Export account tracking JSON to Excel.")
    parser.add_argument("snapshot_json", help="Path to account snapshot JSON.")
    parser.add_argument(
        "--output",
        default="",
        help="Output xlsx path. Default: same folder/name as JSON.",
    )
    return parser.parse_args()


def load_snapshot(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def style_header(row):
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in row:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def auto_width(ws, max_width=60):
    for column_cells in ws.columns:
        max_len = 0
        col = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(value), max_width))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.column_dimensions[col].width = max(12, min(max_len + 2, max_width))


def write_profile_sheet(wb, snapshot):
    ws = wb.active
    ws.title = "账号概览"
    profile = snapshot.get("profile", {})
    totals = snapshot.get("totals", {})
    rows = [
        ["字段", "值"],
        ["追踪时间", snapshot.get("tracked_at", "")],
        ["主页链接", snapshot.get("user_url", "")],
        ["昵称", profile.get("nickname", "")],
        ["抖音号", profile.get("unique_id", "")],
        ["UID", profile.get("uid", "")],
        ["Sec UID", profile.get("sec_uid", "")],
        ["简介", profile.get("signature", "")],
        ["IP 属地", profile.get("ip_location", "")],
        ["粉丝数", profile.get("follower_count", 0)],
        ["关注数", profile.get("following_count", 0)],
        ["作品数", profile.get("aweme_count", 0)],
        ["获赞/收藏总数", profile.get("total_favorited", 0)],
        ["本次返回作品数", totals.get("work_count_returned", 0)],
        ["本次作品点赞合计", totals.get("digg_count_sum", 0)],
        ["本次作品评论合计", totals.get("comment_count_sum", 0)],
        ["本次作品收藏合计", totals.get("collect_count_sum", 0)],
        ["本次作品分享合计", totals.get("share_count_sum", 0)],
    ]
    for row in rows:
        ws.append(row)
    style_header(ws[1])
    auto_width(ws)


def write_works_sheet(wb, snapshot):
    ws = wb.create_sheet("作品明细")
    headers = [
        "序号",
        "作品ID",
        "发布时间",
        "标题/文案",
        "点赞",
        "评论",
        "收藏",
        "分享",
        "话题",
        "作品链接",
    ]
    ws.append(headers)
    style_header(ws[1])

    for index, work in enumerate(snapshot.get("works", []), 1):
        ws.append(
            [
                index,
                work.get("work_id", ""),
                work.get("create_time_text", ""),
                work.get("title") or work.get("desc", ""),
                work.get("digg_count", 0),
                work.get("comment_count", 0),
                work.get("collect_count", 0),
                work.get("share_count", 0),
                ", ".join(work.get("topics", [])),
                work.get("work_url", ""),
            ]
        )
    ws.freeze_panes = "A2"
    auto_width(ws)


def export_excel(snapshot, output_path):
    wb = openpyxl.Workbook()
    write_profile_sheet(wb, snapshot)
    write_works_sheet(wb, snapshot)
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)


def main():
    args = parse_args()
    snapshot_path = os.path.abspath(args.snapshot_json)
    snapshot = load_snapshot(snapshot_path)
    output_path = args.output
    if not output_path:
        output_path = os.path.splitext(snapshot_path)[0] + ".xlsx"
    output_path = os.path.abspath(output_path)
    export_excel(snapshot, output_path)
    print(f"Excel exported: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
